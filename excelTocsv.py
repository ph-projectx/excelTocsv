import openpyxl, os, csv
wb = openpyxl.Workbook()


for excelFile in os.listdir('.'):
    #skip non-xlsx files, load the workbook object.
    if not excelFile.endswith('.xlsx'):
        continue
    wb = openpyxl.load_workbook(excelFile)
    for sheetName in wb.sheetnames:
        #loop thourgh every sheet in the workbook.
        sheet = wb[sheetName]
        # Create the CSV filename from the Excel filename and sheet title.
        if '.' in excelFile:
            excel_name = '.'.join(excelFile.split('.')[:-1])
            new_csv = excel_name +'_'+ sheetName +'.csv'
            print(new_csv)
        # Create the csv.writer object for this CSV file.
        csvFile = open(new_csv, 'w', newline='')
        csvWriter = csv.writer(csvFile)

        # Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            rowDATA = [] # append each cell to the list
            #loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):
                data = sheet.cell(row=rowNum, column=colNum).value
                rowDATA.append(data)

            #write the rowDATA list to the csv file.
            csvWriter.writerow(rowDATA)
        csvFile.close()